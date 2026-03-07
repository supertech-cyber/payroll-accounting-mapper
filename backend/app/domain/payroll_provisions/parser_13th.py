from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.payroll_provisions.models import (
    Provision13thCostCenterSnapshot,
    Provision13thEntry,
    Provision13thResult,
)


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value)


def to_float(value: Any) -> float:
    if value in (None, "", " "):
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    return float(Decimal(text))


def parse_company(text: str) -> tuple[str, str]:
    match = re.match(r"Empresa:\s*(\d+)\s*-\s*(.+)$", text)
    if not match:
        return "", text
    return match.group(1), match.group(2).strip()


def parse_cnpj(text: str) -> tuple[str | None, str | None]:
    match = re.search(r"CNPJ[:\s]*([\d\.\-\/]+)", text, flags=re.IGNORECASE)
    if not match:
        return None, None

    cnpj = only_digits(match.group(1))
    if len(cnpj) != 14:
        return None, None

    return cnpj, cnpj[:8]


def find_company_cnpj(ws, company_row: int) -> tuple[str | None, str | None]:
    max_cols_to_scan = min(ws.max_column, 54)

    for row in range(company_row, min(company_row + 2, ws.max_row) + 1):
        for col in range(1, max_cols_to_scan + 1):
            text = norm(ws.cell(row, col).value)
            if "CNPJ" in text.upper():
                cnpj, cnpj_base = parse_cnpj(text)
                if cnpj:
                    return cnpj, cnpj_base

    return None, None


def parse_competence(text: str) -> str:
    """
    Exemplo:
    'Relatório de provisão de 13º salário 09/2024' -> '2024-09'
    """
    text = text.upper()
    match = re.search(r"(\d{2})/(\d{4})", text)
    if not match:
        raise ValueError(f"Competência não encontrada em: {text}")
    month = match.group(1)
    year = match.group(2)
    return f"{year}-{month}"


def competence_to_display(competence: str) -> str:
    """
    '2024-09' -> '09/2024'
    """
    year, month = competence.split("-")
    return f"{month}/{year}"


def is_header_row(text: str) -> bool:
    return text.upper().startswith("RELATÓRIO DE PROVISÃO DE 13")


def is_company_row(text: str) -> bool:
    return text.startswith("Empresa:")


def is_cost_center_row(text: str) -> bool:
    return text.startswith("TOTAL CENTRO DE CUSTO:")


def parse_cost_center(text: str) -> tuple[str, str]:
    """
    'TOTAL CENTRO DE CUSTO: 1 - Administrativo'
    """
    match = re.match(r"TOTAL CENTRO DE CUSTO:\s*(\d+)\s*-\s*(.+)$", text)
    if not match:
        return "", text
    return match.group(1), match.group(2).strip()


def parse_single_13th_report(path: str | Path) -> list[Provision13thCostCenterSnapshot]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    company_code = ""
    company_name = ""
    company_cnpj: str | None = None
    company_cnpj_base: str | None = None
    competence = ""

    snapshots: list[Provision13thCostCenterSnapshot] = []

    row = 1
    while row <= ws.max_row:
        col_a = norm(ws.cell(row, 1).value)

        if is_header_row(col_a):
            competence = parse_competence(col_a)
            row += 1
            continue

        if is_company_row(col_a):
            company_code, company_name = parse_company(col_a)
            company_cnpj, company_cnpj_base = find_company_cnpj(ws, row)
            row += 1
            continue

        if is_cost_center_row(col_a):
            cost_center_code, cost_center_name = parse_cost_center(col_a)

            # Procurar a linha "Total saldo:" logo abaixo
            scan_row = row + 1
            total_saldo_row = None

            while scan_row <= ws.max_row:
                label = norm(ws.cell(scan_row, 1).value)

                if (
                    is_cost_center_row(label)
                    or is_header_row(label)
                    or is_company_row(label)
                ):
                    break

                if label == "Total saldo:":
                    total_saldo_row = scan_row
                    break

                scan_row += 1

            if total_saldo_row is not None:
                snapshots.append(
                    Provision13thCostCenterSnapshot(
                        company_code=company_code,
                        company_name=company_name,
                        company_cnpj=company_cnpj,
                        company_cnpj_base=company_cnpj_base,
                        competence=competence,
                        cost_center_code=cost_center_code,
                        cost_center_name=cost_center_name,
                        total_saldo_13th=to_float(ws.cell(total_saldo_row, 2).value),
                        total_saldo_fgts=to_float(ws.cell(total_saldo_row, 7).value),
                        total_saldo_inss=to_float(ws.cell(total_saldo_row, 9).value),
                        total_saldo_terc=to_float(ws.cell(total_saldo_row, 12).value),
                        total_saldo_rat=to_float(ws.cell(total_saldo_row, 15).value),
                    )
                )

            row = scan_row
            continue

        row += 1

    return snapshots


def build_snapshot_index(
    snapshots: list[Provision13thCostCenterSnapshot],
) -> dict[tuple[str, str], Provision13thCostCenterSnapshot]:
    index: dict[tuple[str, str], Provision13thCostCenterSnapshot] = {}
    for item in snapshots:
        key = (item.company_code, item.cost_center_code)
        index[key] = item
    return index


def parse_13th_reports(
    report_a_path: str | Path,
    report_b_path: str | Path,
) -> list[Provision13thResult]:
    snapshots_a = parse_single_13th_report(report_a_path)
    snapshots_b = parse_single_13th_report(report_b_path)

    if not snapshots_a or not snapshots_b:
        raise ValueError(
            "Não foi possível ler os dados de um dos relatórios de provisão de 13º."
        )

    competence_a = snapshots_a[0].competence
    competence_b = snapshots_b[0].competence

    if competence_a == competence_b:
        raise ValueError(
            "Os dois relatórios possuem a mesma competência. É necessário enviar competências diferentes."
        )

    if competence_a < competence_b:
        previous_snapshots = snapshots_a
        current_snapshots = snapshots_b
        competence_previous = competence_a
        competence_current = competence_b
    else:
        previous_snapshots = snapshots_b
        current_snapshots = snapshots_a
        competence_previous = competence_b
        competence_current = competence_a

    previous_index = build_snapshot_index(previous_snapshots)
    current_index = build_snapshot_index(current_snapshots)

    all_keys = sorted(set(previous_index.keys()) | set(current_index.keys()))

    results: list[Provision13thResult] = []

    for key in all_keys:
        prev = previous_index.get(key)
        curr = current_index.get(key)

        base = curr or prev
        if base is None:
            continue

        prev_13 = prev.total_saldo_13th if prev else 0.0
        curr_13 = curr.total_saldo_13th if curr else 0.0

        prev_fgts = prev.total_saldo_fgts if prev else 0.0
        curr_fgts = curr.total_saldo_fgts if curr else 0.0

        prev_inss_total = (
            (prev.total_saldo_inss + prev.total_saldo_terc + prev.total_saldo_rat)
            if prev
            else 0.0
        )
        curr_inss_total = (
            (curr.total_saldo_inss + curr.total_saldo_terc + curr.total_saldo_rat)
            if curr
            else 0.0
        )

        results.append(
            Provision13thResult(
                company_code=base.company_code,
                company_name=base.company_name,
                company_cnpj=base.company_cnpj,
                company_cnpj_base=base.company_cnpj_base,
                competence_previous=competence_previous,
                competence_current=competence_current,
                cost_center_code=base.cost_center_code,
                cost_center_name=base.cost_center_name,
                entries=[
                    Provision13thEntry(
                        entry_code="PROV13",
                        entry_description=f"PROV13 {competence_to_display(competence_current)}",
                        amount_previous=round(prev_13, 2),
                        amount_current=round(curr_13, 2),
                        amount_difference=round(curr_13 - prev_13, 2),
                    ),
                    Provision13thEntry(
                        entry_code="PROVFGTS13",
                        entry_description=f"PROVFGTS13 {competence_to_display(competence_current)}",
                        amount_previous=round(prev_fgts, 2),
                        amount_current=round(curr_fgts, 2),
                        amount_difference=round(curr_fgts - prev_fgts, 2),
                    ),
                    Provision13thEntry(
                        entry_code="PROVINSS13",
                        entry_description=f"PROVINSS13 {competence_to_display(competence_current)}",
                        amount_previous=round(prev_inss_total, 2),
                        amount_current=round(curr_inss_total, 2),
                        amount_difference=round(curr_inss_total - prev_inss_total, 2),
                    ),
                ],
            )
        )

    return results
