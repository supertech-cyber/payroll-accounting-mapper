from __future__ import annotations

import re
from decimal import Decimal
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.payroll_mirror.models import EventItem, PayrollBlock


def norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def only_digits(value: str) -> str:
    return re.sub(r"\D", "", value)


def to_float(value: Any) -> float:
    if value in (None, "", " "):
        return 0.0

    if isinstance(value, (int, float)):
        return float(value)

    text = str(value).strip()
    if not text:
        return 0.0

    if "," in text and "." in text:
        text = text.replace(".", "").replace(",", ".")
    elif "," in text:
        text = text.replace(",", ".")

    return float(Decimal(text))


def parse_company(text: str) -> tuple[str, str]:
    match = re.match(r"Empresa:\s*(\d+)\s*-\s*(.+)$", text)
    if not match:
        return "", text
    return match.group(1), match.group(2).strip()


def parse_cnpj(text: str) -> tuple[str | None, str | None]:
    """
    Exemplo:
    'Santa Maria/RS - CNPJ:90.146.531/0001-80'
    -> ('90146531000180', '90146531')
    """
    match = re.search(r"CNPJ[:\s]*([\d\.\-\/]+)", text, flags=re.IGNORECASE)
    if not match:
        return None, None

    cnpj = only_digits(match.group(1))
    if len(cnpj) != 14:
        return None, None

    return cnpj, cnpj[:8]


def find_company_cnpj(ws, company_row: int) -> tuple[str | None, str | None]:
    """
    Procura o CNPJ perto da linha da empresa, porque ele pode estar:
    - na mesma linha, bem à direita
    - na linha seguinte
    """
    max_cols_to_scan = min(ws.max_column, 54)

    for row in range(company_row, min(company_row + 2, ws.max_row) + 1):
        for col in range(1, max_cols_to_scan + 1):
            text = norm(ws.cell(row, col).value)
            if "CNPJ" in text.upper():
                cnpj, cnpj_base = parse_cnpj(text)
                if cnpj:
                    return cnpj, cnpj_base

    return None, None


def parse_competence(text: str) -> str:
    """
    Exemplo:
    'Espelho e resumo da folha mensal referente ao mês de JANEIRO/2026'
    -> '2026-01'
    """
    text = text.upper()

    months = {
        "JANEIRO": "01",
        "FEVEREIRO": "02",
        "MARÇO": "03",
        "MARCO": "03",
        "ABRIL": "04",
        "MAIO": "05",
        "JUNHO": "06",
        "JULHO": "07",
        "AGOSTO": "08",
        "SETEMBRO": "09",
        "OUTUBRO": "10",
        "NOVEMBRO": "11",
        "DEZEMBRO": "12",
    }

    match = re.search(r"M[ÊE]S DE\s+([A-ZÇ]+)\s*/\s*(\d{4})", text)
    if not match:
        raise ValueError(f"Competência não encontrada em: {text}")

    month_name = match.group(1)
    year = match.group(2)
    month = months[month_name]
    return f"{year}-{month}"


def parse_cost_center(text: str) -> tuple[str | None, str | None]:
    match = re.match(r"(\d+)\s*-\s*(.+)$", text)
    if not match:
        return None, None
    return match.group(1), match.group(2).strip()


def is_page_header(text: str) -> bool:
    return text.upper().startswith("ESPELHO")


def is_company_row(text: str) -> bool:
    return text.startswith("Empresa:")


def is_cost_center_label(text: str) -> bool:
    return text == "Centro de Custo :"


def is_summary_row(text: str) -> bool:
    return text == "RESUMO GERAL"


def is_gps_row(text: str) -> bool:
    return text.startswith("Analítico GPS")


def is_events_header(left: str, right: str) -> bool:
    return left.startswith("PROVENTOS") or right.startswith("DESCONTOS")


def parse_gps_patronal_value(raw: str) -> float | None:
    """
    Exemplo:
    '26.001,88 (Bruto) - 6.394,21 (Segurados)'
    => 19607.67
    """
    if not raw:
        return None

    values = re.findall(r"[\d\.\,]+", raw)
    if len(values) < 2:
        return None

    bruto = to_float(values[0])
    segurados = to_float(values[1])
    return round(bruto - segurados, 2)


def extract_event_from_row(ws, row: int) -> list[EventItem]:
    items: list[EventItem] = []

    # PROVENTOS
    prov_code = norm(ws.cell(row, 1).value)  # A
    prov_desc = norm(ws.cell(row, 2).value)  # B
    prov_amount = ws.cell(row, 21).value  # U

    if prov_code and prov_desc and prov_code not in {"PROVENTOS", "RESUMO GERAL"}:
        amount = to_float(prov_amount)
        if amount != 0:
            items.append(
                EventItem(
                    entry_type="PROVENTO",
                    event_code=prov_code,
                    description=prov_desc,
                    amount=amount,
                )
            )

    # DESCONTOS
    disc_code = norm(ws.cell(row, 27).value)  # AA
    disc_desc = norm(ws.cell(row, 30).value)  # AD
    disc_amount = ws.cell(row, 52).value  # AZ

    if disc_code and disc_desc and disc_code != "DESCONTOS":
        amount = to_float(disc_amount)
        if amount != 0:
            items.append(
                EventItem(
                    entry_type="DESCONTO",
                    event_code=disc_code,
                    description=disc_desc,
                    amount=amount,
                )
            )

    return items


def read_summary(ws, start_row: int) -> tuple[dict[str, float], int]:
    summary: dict[str, float] = {}
    row = start_row + 1

    while row <= ws.max_row:
        label = norm(ws.cell(row, 1).value)

        if not label:
            row += 1
            continue

        if (
            label.startswith("Analítico GPS")
            or is_page_header(label)
            or is_company_row(label)
        ):
            break

        value = ws.cell(row, 9).value

        if label not in {"RESUMO GERAL"}:
            try:
                if value not in (None, "", " "):
                    summary[label] = to_float(value)
            except Exception:
                pass

        row += 1

    return summary, row


def read_gps(ws, start_row: int) -> tuple[dict[str, str | float], int]:
    gps: dict[str, str | float] = {}
    row = start_row

    while row <= ws.max_row:
        label = norm(ws.cell(row, 1).value)

        if not label:
            row += 1
            continue

        if (is_page_header(label) or is_company_row(label)) and row != start_row:
            break

        if label.startswith("GPS - >"):
            gps["gps_raw"] = norm(ws.cell(row, 3).value)

        if label.startswith("GPS patronal - >"):
            raw_value = norm(ws.cell(row, 7).value)
            gps["gps_patronal_raw"] = raw_value
            parsed_value = parse_gps_patronal_value(raw_value)
            if parsed_value is not None:
                gps["gps_patronal_value"] = parsed_value

        row += 1

    return gps, row


def parse_payroll_mirror(path: str | Path) -> list[PayrollBlock]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    blocks: list[PayrollBlock] = []

    current_company_code = ""
    current_company_name = ""
    current_company_cnpj: str | None = None
    current_company_cnpj_base: str | None = None
    current_competence = ""
    current_cost_center_code: str | None = None
    current_cost_center_name: str | None = None
    current_block: PayrollBlock | None = None

    row = 1
    while row <= ws.max_row:
        col_a = norm(ws.cell(row, 1).value)
        col_h = norm(ws.cell(row, 8).value)
        col_z = norm(ws.cell(row, 26).value)

        if is_page_header(col_a):
            current_competence = parse_competence(col_a)
            row += 1
            continue

        if is_company_row(col_a):
            current_company_code, current_company_name = parse_company(col_a)

            cnpj, cnpj_base = find_company_cnpj(ws, row)
            current_company_cnpj = cnpj
            current_company_cnpj_base = cnpj_base

            row += 1
            continue

        if is_cost_center_label(col_a):
            current_cost_center_code, current_cost_center_name = parse_cost_center(
                col_h
            )
            row += 1
            continue

        if is_events_header(col_a, col_z):
            is_totalizer = (
                current_cost_center_code is None and current_cost_center_name is None
            )

            current_block = PayrollBlock(
                company_code=current_company_code,
                company_name=current_company_name,
                company_cnpj=current_company_cnpj,
                company_cnpj_base=current_company_cnpj_base,
                competence=current_competence,
                cost_center_code=current_cost_center_code,
                cost_center_name=current_cost_center_name,
                is_totalizer=is_totalizer,
                source_start_row=row,
            )
            blocks.append(current_block)

            row += 1
            while row <= ws.max_row:
                a = norm(ws.cell(row, 1).value)

                if (
                    is_summary_row(a)
                    or is_gps_row(a)
                    or is_page_header(a)
                    or is_company_row(a)
                ):
                    break

                current_block.events.extend(extract_event_from_row(ws, row))
                row += 1

            continue

        if is_summary_row(col_a) and current_block:
            summary, new_row = read_summary(ws, row)
            current_block.summary = summary
            row = new_row
            continue

        if is_gps_row(col_a) and current_block:
            gps, new_row = read_gps(ws, row)
            current_block.gps = gps

            # próximo bloco decide se tem centro de custo ou se é totalizador
            current_cost_center_code = None
            current_cost_center_name = None

            row = new_row
            continue

        row += 1

    return blocks
