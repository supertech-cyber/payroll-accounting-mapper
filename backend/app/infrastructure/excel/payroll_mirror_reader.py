from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.domain.payroll_mirror.models import EventItem, PayrollBlock
from app.infrastructure.excel._shared import (
    find_company_cnpj,
    norm,
    parse_company,
    to_float,
)


def _parse_competence(text: str) -> str:
    """
    Example:
    'Espelho e resumo da folha mensal referente ao mês de JANEIRO/2026'
    -> '2026-01'
    """
    text = text.upper()
    months = {
        "JANEIRO": "01",
        "FEVEREIRO": "02",
        "MARÇO": "03",
        "MARCO": "03",
        "ABRIL": "04",
        "MAIO": "05",
        "JUNHO": "06",
        "JULHO": "07",
        "AGOSTO": "08",
        "SETEMBRO": "09",
        "OUTUBRO": "10",
        "NOVEMBRO": "11",
        "DEZEMBRO": "12",
    }
    match = re.search(r"M[ÊE]S DE\s+([A-ZÇ]+)\s*/\s*(\d{4})", text)
    if not match:
        raise ValueError(f"Competência não encontrada em: {text}")
    month = months[match.group(1)]
    return f"{match.group(2)}-{month}"


def _parse_cost_center(text: str) -> tuple[str, str]:
    """Extract (code, name) from the CC label next to 'Centro de Custo :'.

    Numeric layout: '1 - Administrativo'  → ('1', 'Administrativo')
    Text-only:      'Colaboradores sem centro de custo'
                    '(Empresa) Agefer Comercio e Cereais LTDA'
                    → use full text as both code and name.
    """
    match = re.match(r"(\d+)\s*-\s*(.+)$", text)
    if not match:
        return text, text
    return match.group(1), match.group(2).strip()


def _is_page_header(text: str) -> bool:
    return text.upper().startswith("ESPELHO")


def _is_company_row(text: str) -> bool:
    return text.startswith("Empresa:")


def _is_cost_center_label(text: str) -> bool:
    return text == "Centro de Custo :"


def _is_summary_row(text: str) -> bool:
    return text == "RESUMO GERAL"


def _is_gps_row(text: str) -> bool:
    return text.startswith("Analítico GPS")


def _is_events_header(left: str, right: str) -> bool:
    return left.startswith("PROVENTOS") or right.startswith("DESCONTOS")


def _find_descontos_col(ws: Any, header_row: int) -> int:
    """
    Scan the header row to find the column that contains 'DESCONTOS'.
    Different file generations may shift the layout by 1+ columns.
    Falls back to column 27 (AA) if not found.
    """
    for col in range(1, ws.max_column + 1):
        v = norm(ws.cell(header_row, col).value)
        if v.startswith("DESCONTOS"):
            return col
    return 27  # original / fallback layout


def _parse_gps_patronal_value(raw: str) -> float | None:
    """
    Example:
    '26.001,88 (Bruto) - 6.394,21 (Segurados)'
    -> 19607.67
    """
    if not raw:
        return None
    values = re.findall(r"[\d\.\,]+", raw)
    if len(values) < 2:
        return None
    return round(to_float(values[0]) - to_float(values[1]), 2)


def _extract_event_from_row(
    ws: Any,
    row: int,
    descontos_col: int,
    disc_amount_col: int,
) -> list[EventItem]:
    """
    Extract provento and desconto items from a single data row.

    Column layout is derived dynamically:
      - prov_code / prov_desc: always cols A / B (fixed)
      - prov_amount: descontos_col - 6
      - disc_code:   descontos_col  (same column as DESCONTOS header)
      - disc_desc:   descontos_col + 3
      - disc_amount: disc_amount_col  (= ws.max_column - 2)
    """
    items: list[EventItem] = []

    prov_code = norm(ws.cell(row, 1).value)  # col A — fixed
    prov_desc = norm(ws.cell(row, 2).value)  # col B — fixed
    prov_amount = ws.cell(row, descontos_col - 6).value  # relative to DESCONTOS
    if prov_code and prov_desc and prov_code not in {"PROVENTOS", "RESUMO GERAL"}:
        amount = to_float(prov_amount)
        if amount != 0:
            items.append(
                EventItem(
                    entry_type="PROVENTO",
                    event_code=prov_code,
                    description=prov_desc,
                    amount=amount,
                )
            )

    disc_code = norm(ws.cell(row, descontos_col).value)  # same col as DESCONTOS header
    disc_desc = norm(ws.cell(row, descontos_col + 3).value)  # 3 cols after DESCONTOS
    disc_amount = ws.cell(row, disc_amount_col).value  # max_column - 2
    if disc_code and disc_desc and disc_code != "DESCONTOS":
        amount = to_float(disc_amount)
        if amount != 0:
            items.append(
                EventItem(
                    entry_type="DESCONTO",
                    event_code=disc_code,
                    description=disc_desc,
                    amount=amount,
                )
            )

    return items


def _read_summary(ws: Any, start_row: int) -> tuple[dict[str, float], int]:
    summary: dict[str, float] = {}
    row = start_row + 1
    while row <= ws.max_row:
        label = norm(ws.cell(row, 1).value)
        if not label:
            row += 1
            continue
        if (
            label.startswith("Analítico GPS")
            or _is_page_header(label)
            or _is_company_row(label)
        ):
            break
        value = ws.cell(row, 9).value
        if label != "RESUMO GERAL":
            try:
                if value not in (None, "", " "):
                    summary[label] = to_float(value)
            except Exception:
                pass
        row += 1
    return summary, row


def _read_gps(ws: Any, start_row: int) -> tuple[dict[str, str | float], int]:
    gps: dict[str, str | float] = {}
    row = start_row
    while row <= ws.max_row:
        label = norm(ws.cell(row, 1).value)
        if not label:
            row += 1
            continue
        if (_is_page_header(label) or _is_company_row(label)) and row != start_row:
            break
        if label.startswith("GPS - >"):
            gps["gps_raw"] = norm(ws.cell(row, 3).value)
        if label.startswith("GPS patronal - >"):
            raw_value = norm(ws.cell(row, 7).value)
            gps["gps_patronal_raw"] = raw_value
            parsed = _parse_gps_patronal_value(raw_value)
            if parsed is not None:
                gps["gps_patronal_value"] = parsed
        row += 1
    return gps, row


def parse_payroll_mirror(path: str | Path) -> list[PayrollBlock]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    blocks: list[PayrollBlock] = []
    current_company_code = ""
    current_company_name = ""
    current_company_cnpj: str | None = None
    current_company_cnpj_base: str | None = None
    current_competence = ""
    current_cost_center_code: str | None = None
    current_cost_center_name: str | None = None
    current_block: PayrollBlock | None = None

    row = 1
    while row <= ws.max_row:
        col_a = norm(ws.cell(row, 1).value)
        col_h = norm(ws.cell(row, 8).value)
        col_z = norm(ws.cell(row, 26).value)

        if _is_page_header(col_a):
            current_competence = _parse_competence(col_a)
            row += 1
            continue

        if _is_company_row(col_a):
            current_company_code, current_company_name = parse_company(col_a)
            current_company_cnpj, current_company_cnpj_base = find_company_cnpj(ws, row)
            row += 1
            continue

        if _is_cost_center_label(col_a):
            current_cost_center_code, current_cost_center_name = _parse_cost_center(
                col_h
            )
            row += 1
            continue

        if _is_events_header(col_a, col_z):
            # Detect column layout for this block — different file generations
            # may have the DESCONTOS section shifted by one or more columns.
            descontos_col = _find_descontos_col(ws, row)
            disc_amount_col = ws.max_column - 2

            is_totalizer = (
                current_cost_center_code is None and current_cost_center_name is None
            )
            current_block = PayrollBlock(
                company_code=current_company_code,
                company_name=current_company_name,
                company_cnpj=current_company_cnpj,
                company_cnpj_base=current_company_cnpj_base,
                competence=current_competence,
                cost_center_code=current_cost_center_code,
                cost_center_name=current_cost_center_name,
                is_totalizer=is_totalizer,
                source_start_row=row,
            )
            blocks.append(current_block)
            row += 1
            while row <= ws.max_row:
                a = norm(ws.cell(row, 1).value)
                if (
                    _is_summary_row(a)
                    or _is_gps_row(a)
                    or _is_page_header(a)
                    or _is_company_row(a)
                ):
                    break
                current_block.events.extend(
                    _extract_event_from_row(ws, row, descontos_col, disc_amount_col)
                )
                row += 1
            continue

        if _is_summary_row(col_a) and current_block:
            current_block.summary, row = _read_summary(ws, row)
            continue

        if _is_gps_row(col_a) and current_block:
            current_block.gps, row = _read_gps(ws, row)
            current_cost_center_code = None
            current_cost_center_name = None
            continue

        row += 1

    return blocks
