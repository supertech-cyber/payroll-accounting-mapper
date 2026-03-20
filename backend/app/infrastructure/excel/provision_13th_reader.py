from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.infrastructure.excel._shared import (
    find_company_cnpj,
    norm,
    parse_company,
    to_float,
)


def _parse_competence(text: str) -> str:
    text = text.upper()
    match = re.search(r"(\d{2})/(\d{4})", text)
    if not match:
        raise ValueError(f"Competência não encontrada em: {text}")
    return f"{match.group(2)}-{match.group(1)}"


def _is_header_row(text: str) -> bool:
    return "RELATÓRIO DE PROVISÃO DE 13" in text.upper()


def _is_company_row(text: str) -> bool:
    return text.startswith("Empresa:")


def _is_cost_center_row(text: str) -> bool:
    return text.startswith("TOTAL CENTRO DE CUSTO:")


def _parse_cost_center(text: str) -> tuple[str, str]:
    """Extract (code, name) from 'TOTAL CENTRO DE CUSTO: ...' lines.

    Numeric: 'TOTAL CENTRO DE CUSTO: 1 - Administrativo' → ('1', 'Administrativo')
    Text:    'TOTAL CENTRO DE CUSTO: (Empresa) Agefer Comercio e Cereais LTDA'
             → ('(Empresa) Agefer Comercio e Cereais LTDA', same)
    """
    match = re.match(r"TOTAL CENTRO DE CUSTO:\s*(\d+)\s*-\s*(.+)$", text)
    if not match:
        m2 = re.match(r"TOTAL CENTRO DE CUSTO:\s*(.+)$", text, re.IGNORECASE)
        rest = m2.group(1).strip() if m2 else text
        return rest, rest
    return match.group(1), match.group(2).strip()


@dataclass
class _13thSnapshot:
    """Internal reading artefact — not a domain entity."""

    company_code: str
    company_name: str
    company_cnpj: str | None
    company_cnpj_base: str | None
    competence: str
    cost_center_code: str
    cost_center_name: str
    total_saldo_13th: float
    total_saldo_fgts: float
    total_saldo_inss: float
    total_saldo_terc: float
    total_saldo_rat: float


def parse_single_13th_report(path: str | Path) -> list[_13thSnapshot]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    company_code = ""
    company_name = ""
    company_cnpj: str | None = None
    company_cnpj_base: str | None = None
    competence = ""
    snapshots: list[_13thSnapshot] = []

    row = 1
    while row <= ws.max_row:
        col_a = norm(ws.cell(row, 1).value)

        if _is_header_row(col_a):
            competence = _parse_competence(col_a)
            row += 1
            continue

        if _is_company_row(col_a):
            company_code, company_name = parse_company(col_a)
            company_cnpj, company_cnpj_base = find_company_cnpj(ws, row)
            row += 1
            continue

        if _is_cost_center_row(col_a):
            cost_center_code, cost_center_name = _parse_cost_center(col_a)
            scan_row = row + 1
            total_saldo_row = None
            while scan_row <= ws.max_row:
                label = norm(ws.cell(scan_row, 1).value)
                if (
                    _is_cost_center_row(label)
                    or _is_header_row(label)
                    or _is_company_row(label)
                ):
                    break
                if label == "Total saldo:":
                    total_saldo_row = scan_row
                    break
                scan_row += 1
            if total_saldo_row is not None:
                snapshots.append(
                    _13thSnapshot(
                        company_code=company_code,
                        company_name=company_name,
                        company_cnpj=company_cnpj,
                        company_cnpj_base=company_cnpj_base,
                        competence=competence,
                        cost_center_code=cost_center_code,
                        cost_center_name=cost_center_name,
                        total_saldo_13th=to_float(ws.cell(total_saldo_row, 3).value),
                        total_saldo_fgts=to_float(ws.cell(total_saldo_row, 8).value),
                        total_saldo_inss=to_float(ws.cell(total_saldo_row, 10).value),
                        total_saldo_terc=to_float(ws.cell(total_saldo_row, 13).value),
                        total_saldo_rat=to_float(ws.cell(total_saldo_row, 16).value),
                    )
                )
            row = scan_row
            continue

        row += 1

    return snapshots
