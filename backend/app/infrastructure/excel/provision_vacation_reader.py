from __future__ import annotations

import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from app.infrastructure.excel._shared import (
    find_company_cnpj,
    norm,
    parse_company,
    to_float,
)


def _parse_competence(text: str) -> str:
    text = text.upper()
    match = re.search(r"(\d{2})/(\d{4})", text)
    if not match:
        raise ValueError(f"Competência não encontrada em: {text}")
    return f"{match.group(2)}-{match.group(1)}"


def _is_header_row(text: str) -> bool:
    u = text.upper()
    return (
        "RELATÓRIO DE PROVISÃO DE FÉRIAS" in u or "RELATORIO DE PROVISAO DE FERIAS" in u
    )


def _is_company_row(text: str) -> bool:
    return text.startswith("Empresa:")


def _is_cost_center_row(text: str) -> bool:
    return text.startswith("TOTAL CENTRO DE CUSTO")


def _parse_cost_center(text: str) -> tuple[str, str]:
    match = re.match(r"TOTAL CENTRO DE CUSTO\s*:\s*(\d+)\s*-\s*(.+)$", text)
    if not match:
        return "", text
    return match.group(1), match.group(2).strip()


@dataclass
class _VacationSnapshot:
    """Internal reading artefact — not a domain entity."""

    company_code: str
    company_name: str
    company_cnpj: str | None
    company_cnpj_base: str | None
    competence: str
    cost_center_code: str
    cost_center_name: str
    total_saldo_vacation: float
    total_saldo_bonus: float
    total_saldo_fgts: float
    total_saldo_inss: float
    total_saldo_terc: float
    total_saldo_rat: float


def parse_single_vacation_report(path: str | Path) -> list[_VacationSnapshot]:
    wb = load_workbook(path, data_only=True)
    ws = wb.active

    company_code = ""
    company_name = ""
    company_cnpj: str | None = None
    company_cnpj_base: str | None = None
    competence = ""
    snapshots: list[_VacationSnapshot] = []

    row = 1
    while row <= ws.max_row:
        col_a = norm(ws.cell(row, 1).value)

        if _is_header_row(col_a):
            competence = _parse_competence(col_a)
            row += 1
            continue

        if _is_company_row(col_a):
            company_code, company_name = parse_company(col_a)
            company_cnpj, company_cnpj_base = find_company_cnpj(ws, row)
            row += 1
            continue

        if _is_cost_center_row(col_a):
            cost_center_code, cost_center_name = _parse_cost_center(col_a)
            scan_row = row + 1
            total_saldo_row = None
            while scan_row <= ws.max_row:
                label = norm(ws.cell(scan_row, 1).value)
                if (
                    _is_cost_center_row(label)
                    or _is_header_row(label)
                    or _is_company_row(label)
                ):
                    break
                if label == "Total saldo:":
                    total_saldo_row = scan_row
                    break
                scan_row += 1
            if total_saldo_row is not None:
                snapshots.append(
                    _VacationSnapshot(
                        company_code=company_code,
                        company_name=company_name,
                        company_cnpj=company_cnpj,
                        company_cnpj_base=company_cnpj_base,
                        competence=competence,
                        cost_center_code=cost_center_code,
                        cost_center_name=cost_center_name,
                        total_saldo_vacation=to_float(
                            ws.cell(total_saldo_row, 4).value
                        ),
                        total_saldo_bonus=to_float(ws.cell(total_saldo_row, 7).value),
                        total_saldo_fgts=to_float(ws.cell(total_saldo_row, 9).value),
                        total_saldo_inss=to_float(ws.cell(total_saldo_row, 11).value),
                        total_saldo_terc=to_float(ws.cell(total_saldo_row, 14).value),
                        total_saldo_rat=to_float(ws.cell(total_saldo_row, 17).value),
                    )
                )
            row = scan_row
            continue

        row += 1

    return snapshots
