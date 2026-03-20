#!/usr/bin/env python3
"""
Generate 016_seed_novas_empresas.sql from Excel files.
Reads all 7 company group Excel files and produces idempotent SQL.
Run from repo root or backend/ dir.
"""

import re
import openpyxl
from pathlib import Path

BASE = Path(
    "/home/diogorodrigues/geral/supervisao/payroll-accounting-mapper/novas-empresas-eventos"
)
OUTPUT = Path(
    "/home/diogorodrigues/geral/supervisao/payroll-accounting-mapper/backend/database/seeds/016_seed_novas_empresas.sql"
)

GROUPS = [
    {
        "tag_slug": "agefer",
        "tag_label": "AGEFER",
        "empresa_file": "Agefer/Banco de dados/empresas_Agefer.xlsx",
        "eventos_file": "Agefer/Banco de dados/eventos_Agefer.xlsx",
        "output_template": "fpa-elevor",
    },
    {
        "tag_slug": "arenhart",
        "tag_label": "ARENHART",
        "empresa_file": "Arenhart/Banco de dados/empresas_Arenhart.xlsx",
        "eventos_file": "Arenhart/Banco de dados/eventos_Arenhart.xlsx",
        "output_template": "fpa-elevor",
    },
    {
        "tag_slug": "camsul",
        "tag_label": "CAMSUL",
        "empresa_file": "Camsul/Banco de Dados/empresas_Camsul.xlsx",
        "eventos_file": "Camsul/Banco de Dados/eventos_Camsul.xlsx",
        "output_template": "fpa-elevor",
    },
    {
        "tag_slug": "zamarchi",
        "tag_label": "CEREALISTA ZAMARCHI",
        "empresa_file": "Cerealista Zamarchi/Banco de Dados/empresas_Zamarchi.xlsx",
        "eventos_file": "Cerealista Zamarchi/Banco de Dados/eventos_Zamarchi.xlsx",
        "output_template": "fpa-elevor",
    },
    {
        "tag_slug": "cootranscau",
        "tag_label": "COOTRANSCAU",
        "empresa_file": "Cootranscau/Banco de Dados/empresas_Cootranscau.xlsx",
        "eventos_file": "Cootranscau/Banco de Dados/eventos_Cootranscau.xlsx",
        "output_template": "fpa-elevor",
    },
    {
        "tag_slug": "multifertil",
        "tag_label": "MULTIFERTIL",
        "empresa_file": "Multifertil/Banco de Dados/empresas_multi.xlsx",
        "eventos_file": "Multifertil/Banco de Dados/eventos_multi.xlsx",
        "output_template": "fpa-elevor",
    },
    {
        "tag_slug": "sbrubenich",
        "tag_label": "SB RUBENICH",
        "empresa_file": "SB Rubenich/Banco de Dados/empresas_SB Rubenich.xlsx",
        "eventos_file": "SB Rubenich/Banco de Dados/eventos_SB rubenich.xlsx",
        "output_template": "fpa-elevor",
    },
]


def sql_escape(s):
    """Escape a value for SQL single-quoted string."""
    if s is None:
        return "NULL"
    return "'" + str(s).replace("'", "''") + "'"


def to_str_val(v):
    """Convert cell value to string, handling int/float."""
    if v is None:
        return None
    if isinstance(v, float):
        if v == int(v):
            return str(int(v))
        return str(v)
    if isinstance(v, int):
        return str(v)
    return str(v).strip()


def parse_empresas(path):
    """Returns list of (code, name, fpa_batch)."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    result = []
    for r in range(1, ws.max_row + 1):
        cell = ws.cell(r, 1).value
        if not cell:
            continue
        m = re.match(r"Empresa:\s*(\d+)\s*-\s*(.+)", str(cell).strip())
        if m:
            code = m.group(1)
            name = m.group(2).strip()
            fpa_raw = ws.cell(r, 2).value
            if fpa_raw is None:
                fpa_batch = 1
            else:
                try:
                    fpa_batch = int(float(str(fpa_raw)))
                except ValueError:
                    fpa_batch = 1
            result.append((code, name, fpa_batch))
    return result


def parse_cc_header(header, empresa_counts):
    """Returns (code, name) or None if column should be skipped."""
    if header is None:
        return None
    h = str(header).strip()
    if not h:
        return None

    # "123 - Name" or "123- Name" (hyphen or en-dash)
    m = re.match(r"^(\d+)\s*[-–]\s*(.+)$", h)
    if m:
        return m.group(1), m.group(2).strip()

    # "(Empresa) Name" → use full header text as code
    if h.startswith("(Empresa)"):
        name = h[9:].strip().rstrip(")")
        return h, name

    # "Colaboradores sem centro de custo" variants → use full text as code
    if "sem centro" in h.lower():
        return "Colaboradores sem centro de custo", "Colaboradores sem centro de custo"

    return None  # skip unknown / trailing None columns


def parse_eventos(path):
    """
    Returns (cc_list, event_rows).
    cc_list: list of (cc_code, cc_name)
    event_rows: list of (event_code, description, entry_type, credit_account, [debit_per_cc])
    debit_per_cc is aligned to cc_list; None means no mapping for that CC.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    # CC headers start at column 5 (index 4, 0-based)
    empresa_counts = [0]
    cc_list = []
    cc_col_nums = []  # 1-based column numbers for valid CCs

    for col in range(5, ws.max_column + 1):
        header = ws.cell(1, col).value
        parsed = parse_cc_header(header, empresa_counts)
        if parsed:
            cc_list.append(parsed)
            cc_col_nums.append(col)

    # Use an ordered dict to deduplicate by event_code (last occurrence wins)
    event_dict: dict = {}
    for r in range(2, ws.max_row + 1):
        raw_code = ws.cell(r, 1).value
        if raw_code is None:
            continue
        event_code = to_str_val(raw_code)
        if not event_code:
            continue

        description = str(ws.cell(r, 2).value or "").strip()
        entry_type = str(ws.cell(r, 3).value or "P").strip()
        if entry_type not in ("P", "D"):
            entry_type = "P"

        credit_raw = ws.cell(r, 4).value
        if credit_raw is None:
            # Skip rows with no credit; but if a prior row already has this code
            # with a credit, don't overwrite it
            if event_code not in event_dict:
                continue
            else:
                continue  # always skip no-credit rows

        credit_account = to_str_val(credit_raw)

        debits = []
        for col in cc_col_nums:
            v = ws.cell(r, col).value
            debits.append(to_str_val(v))

        event_dict[event_code] = (
            event_code,
            description,
            entry_type,
            credit_account,
            debits,
        )

    event_rows = list(event_dict.values())
    return cc_list, event_rows


def generate_group_sql(group):
    slug = group["tag_slug"]
    label = group["tag_label"]
    template = group.get("output_template", "fpa-elevor")

    empresas = parse_empresas(BASE / group["empresa_file"])
    cc_list, event_rows = parse_eventos(BASE / group["eventos_file"])

    print(
        f"  {label}: {len(empresas)} companies, {len(cc_list)} CCs, {len(event_rows)} events"
    )

    lines = []
    sep = "-- " + "=" * 62
    lines += [
        sep,
        f"-- GROUP: {label}",
        sep,
        "",
    ]

    # ── 1. Tag ────────────────────────────────────────────────────────
    lines += [
        "-- 1. Tag",
        f"INSERT INTO tags (slug, label) VALUES ({sql_escape(slug)}, {sql_escape(label)})",
        "ON CONFLICT (slug) DO UPDATE SET label = EXCLUDED.label;",
        "",
    ]

    # ── 2. Companies ─────────────────────────────────────────────────
    lines.append("-- 2. Companies")
    lines.append(
        "INSERT INTO companies (code, name, output_template, fpa_batch, tag) VALUES"
    )
    for i, (code, name, fpa_batch) in enumerate(empresas):
        comma = "" if i == len(empresas) - 1 else ","
        lines.append(
            f"    ({sql_escape(code)}, {sql_escape(name)}, {sql_escape(template)}, {fpa_batch}, {sql_escape(slug)}){comma}"
        )
    lines += [
        "ON CONFLICT (code) DO UPDATE SET",
        "    name             = EXCLUDED.name,",
        "    output_template  = EXCLUDED.output_template,",
        "    fpa_batch        = EXCLUDED.fpa_batch,",
        "    tag              = EXCLUDED.tag;",
        "",
    ]

    # ── 3. Cost Centers — only for the primary company (fpa_batch=1) ─
    # Secondary companies share CCs via tag-based Priority-2 resolution.
    primary_code = empresas[0][0]  # first entry is always fpa_batch=1
    lines.append("-- 3. Cost Centers")
    lines.append("INSERT INTO cost_centers (code, name, company_id)")
    lines.append("SELECT v.code, v.name, c.id FROM (VALUES")
    for i, (cc_code, cc_name) in enumerate(cc_list):
        comma = "" if i == len(cc_list) - 1 else ","
        lines.append(f"    ({sql_escape(cc_code)}, {sql_escape(cc_name)}){comma}")
    lines += [
        ") AS v(code, name)",
        f"JOIN companies c ON c.code = {sql_escape(primary_code)}",
        "ON CONFLICT (code, company_id) DO UPDATE SET name = EXCLUDED.name;",
        "",
    ]

    # ── 4. Events (global upsert by code) ────────────────────────────
    lines.append(f"-- 4. Events ({len(event_rows)} rows)")
    lines.append("INSERT INTO events (code, description, entry_type) VALUES")
    for i, (ev_code, desc, etype, _, _) in enumerate(event_rows):
        comma = "" if i == len(event_rows) - 1 else ","
        lines.append(
            f"    ({sql_escape(ev_code)}, {sql_escape(desc)}, {sql_escape(etype)}){comma}"
        )
    lines += [
        "ON CONFLICT (code) DO UPDATE",
        "    SET description = EXCLUDED.description,",
        "        entry_type  = EXCLUDED.entry_type;",
        "",
    ]

    # ── 5. Event Mappings — only for the primary company ───────────
    # Secondary companies resolve via tag-based Priority-2 in resolve().
    lines.append("-- 5. Event Mappings")
    mapping_values = []
    for ev_code, _, _, credit, debits in event_rows:
        for idx, (cc_code, _) in enumerate(cc_list):
            debit = debits[idx] if idx < len(debits) else None
            if debit is not None:
                mapping_values.append((ev_code, cc_code, credit, debit))

    if mapping_values:
        lines += [
            f"-- Company {primary_code} ({len(mapping_values)} mappings)",
            "INSERT INTO event_mappings (event_id, cost_center_id, credit_account, debit_account)",
            "SELECT e.id, cc.id, v.credit_account, v.debit_account",
            "FROM (VALUES",
        ]
        for i, (ev_code, cc_code, credit, debit) in enumerate(mapping_values):
            comma = "" if i == len(mapping_values) - 1 else ","
            lines.append(
                f"    ({sql_escape(ev_code)}, {sql_escape(cc_code)}, {sql_escape(credit)}, {sql_escape(debit)}){comma}"
            )
        lines += [
            ") AS v(event_code, cc_code, credit_account, debit_account)",
            "JOIN events       e  ON e.code  = v.event_code",
            "JOIN cost_centers cc ON cc.code = v.cc_code",
            f"    AND cc.company_id = (SELECT id FROM companies WHERE code = {sql_escape(primary_code)})",
            "ON CONFLICT (event_id, cost_center_id) DO UPDATE",
            "    SET credit_account = EXCLUDED.credit_account,",
            "        debit_account  = EXCLUDED.debit_account;",
            "",
        ]

    return lines


def main():
    print("Generating 016_seed_novas_empresas.sql ...")

    all_lines = [
        "-- =================================================================",
        "-- Seed 016: 7 new company groups",
        "--   Agefer, Arenhart, Camsul, Cerealista Zamarchi,",
        "--   Cootranscau, Multifertil, SB Rubenich",
        "-- Generated automatically — safe to re-run (idempotent)",
        "-- =================================================================",
        "",
    ]

    for group in GROUPS:
        all_lines += generate_group_sql(group)
        all_lines.append("")

    # Final verification counts
    all_lines += [
        "-- =================================================================",
        "-- Verification",
        "-- =================================================================",
        "SELECT 'tags'          AS tbl, COUNT(*) FROM tags          UNION ALL",
        "SELECT 'companies',           COUNT(*) FROM companies       UNION ALL",
        "SELECT 'cost_centers',        COUNT(*) FROM cost_centers    UNION ALL",
        "SELECT 'events',              COUNT(*) FROM events          UNION ALL",
        "SELECT 'event_mappings',      COUNT(*) FROM event_mappings;",
    ]

    sql = "\n".join(all_lines)
    OUTPUT.write_text(sql, encoding="utf-8")
    print(f"Written to {OUTPUT}")
    print(f"Total lines: {len(all_lines)}")


if __name__ == "__main__":
    main()
